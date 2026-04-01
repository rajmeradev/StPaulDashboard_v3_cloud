#!/usr/bin/env python3
"""
Excel file reader and parser for St. Paul Plant schedule data — V4.

V4 changes vs V3:
  - MASTER sheet: 2-column layout (col B=Line 1, col C=Line 2).
    scheduleStart is now per-line (B2 / C2). Fat constants at B10/B11/B12.
    SKU database extended to B18:Y58 with Pearson-derived ingredient splits.
  - Schedule sheets: extended to col W (Line 1) / col V (Line 2).
    Q = Raw Req (was "High Fat Req"), S = Cream Req (new), W/V = Segment number.
    "--- SCHEDULE BREAK ---" rows split the schedule into independent segments.
  - Three-ingredient MRP: Raw (Q) + Skim (R) + Cream (S).
  - Liquifier conflict detection: shared resource cross-line check.
  - Segment filter: All / Run N views supported throughout.

Break row timing (Option A):
  On a "--- SCHEDULE BREAK ---" row, if col H contains a datetime value,
  that is the restart time for the next segment. Excel's F formula already
  used this anchor to recalculate K/L/M/N for subsequent rows, so Python
  reads it for display purposes only — no timeline recalculation needed here.

Break detection uses a dual signal:
  - Primary:   Segment col (W for Line 1, V for Line 2) = None → divider row
  - Secondary: Col B exact string "--- SCHEDULE BREAK ---" → explicit isBreakRow flag

CRITICAL LIMITATION (ADR-003):
  openpyxl with data_only=True reads Excel's *cached* formula results.
  If the file was saved without full recalculation, formula cells return None.
  _validate_row() detects this and surfaces it as data quality warnings.
"""

import logging
import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

_LOCAL_TZ = datetime.now().astimezone().tzinfo

import openpyxl

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# SKU Colour Palette — brand-grouped, matches Excel Production Gantt legend
# ---------------------------------------------------------------------------
SKU_COLORS: Dict[str, str] = {
    # ── OV Whole Milk (blue family) ────────────────────────────────────────
    "OV-MILK ORG WH ESL 6-HG CTN BX":  "#AED6F1",
    "OV-MILK ORG WH SP 6-HG CTN BX":   "#AED6F1",
    "OV-MILK ORG WH LAC 6-HG CTN BX":  "#85C1E9",
    "OV-MILK ORG WY-D GRS 6HG CTNBX":  "#7FB3D3",
    # ── OV Skim (green family) ─────────────────────────────────────────────
    "OV-MILK ORG SK 6-HG CTN BX":      "#A9DFBF",
    "OV-MILK ORG SK LAC 6HG CTN BX":   "#7DCEA0",
    "OV-MILK ORG SK 12-QT CTN BX":     "#A9DFBF",
    # ── OV 1% / 2% / Fat Milk (yellow family) ─────────────────────────────
    "OV-MILK ORG 2% 6-HG CTN BX":      "#F9E79F",
    "OV-MILK ORG 2% FT 12-QT CTN BX":  "#F7DC6F",
    "OV-MILK ORG 1% 6-HG CTN BX":      "#FAD7A0",
    "OV-MILK ORG 1% FT 12-QT CTN BX":  "#FAD7A0",
    "OV-MILK ORG 1%CH LAC 6HG CTNBX":  "#FDEBD0",
    "OV-MILK ORG 2% LAC 6-HG CTN BX":  "#F9E79F",
    "OV-MILK ORG WH 3.5% 12QT CTNBX":  "#F5CBA7",
    # ── OV H&H / Cream (orange family) ────────────────────────────────────
    "OV-H&H ORG 12-QTCTN BX":          "#F5CBA7",
    "OV-H&H ORG LAC 12-QT CTN BX":     "#E59866",
    "OV-MILK ORG WY-D CRS 6-59CTNBX":  "#F0B27A",
    # ── TJ Products (indigo/purple family) ────────────────────────────────
    "TJ-H&H ORG 12-PT CTN BX":         "#C39BD3",
    "TJ-H&H ORG 12-QT CTN BX":         "#A569BD",
    "TJ-HVY CRM ORG 12-PT CTN BX":     "#BB8FCE",
    # ── CV Products (teal family) ─────────────────────────────────────────
    "CV-MILK 2% LAC 6-HG CTN BX":      "#76D7C4",
    "CV-MILK WH LAC 6HG CTN BX":       "#48C9B0",
    "CV-H&H 6-QT CTN BX":              "#45B39D",
    "CV-MILK 2% CH LAC 6-59 CTN BX":   "#1ABC9C",
    "CV-NOG UP 6-QT CTN BX":           "#52BE80",
    # ── KEMPS Products (amber family) ─────────────────────────────────────
    "KEMPS-MILK WH LAC 6-HG CTN BX":   "#F8C471",
    "KEMPS-MILK 296 LAC 6-HG CTN BX":  "#F0A500",
    "KEMPS-MILK PUMPKIN 12-QT CTNBX":  "#E67E22",
    "KEMPS-NOG VAN 12-QT CTN BX":      "#CA6F1E",
    "KEMPS-NOG CINN 12-QT CTN BX":     "#BA4A00",
    "KEMPS-MILK PEP MOCHA 12QTCTNBX":  "#935116",
    # ── DNKN Products (mocha/lavender family) ─────────────────────────────
    "DNKN-MILK COFFEE 6-HG CTN BX":    "#D7BDE2",
    "DNKN-MILK CEREAL 6-HG CTN 8X":    "#C39BD3",
    # ── 365 Products (lime family) ────────────────────────────────────────
    "365-MILK ORG SK 12-QT CTN BOX":   "#ABEBC6",
    "365-MILK ORG WH3.5% 12QT CTNBX":  "#82E0AA",
    "365-MILK ORG 2% FT 12-QT CTNBX":  "#58D68D",
    "365-MILK ORG 1% FT 12-QT CTNBX":  "#2ECC71",
    # ── Special types ─────────────────────────────────────────────────────
    "CIP - (Full Cleaning)":            "#F1948A",
    "Downtime":                         "#BDC3C7",
    "Water Flush":                      "#AED6F1",
    "default":                          "#D2B4DE",
}


def get_task_color(sku: str) -> str:
    return SKU_COLORS.get(sku, SKU_COLORS["default"])


def get_task_type(sku: str) -> str:
    upper = sku.upper()
    if upper.strip() == "--- SCHEDULE BREAK ---":
        return "break"
    if "CIP" in upper:
        return "cip"
    if "WATER" in upper or "FLUSH" in upper:
        return "flush"
    if "DOWNTIME" in upper:
        return "downtime"
    return "production"


# ---------------------------------------------------------------------------
# SKU classification helpers
# ---------------------------------------------------------------------------

_SPECIAL_INGREDIENT_KEYWORDS = ("NOG", "PUMPKIN", "MOCHA", "CINN")


def _is_break_row(sku: str) -> bool:
    """Exact match for the schedule break sentinel string."""
    return sku.strip() == "--- SCHEDULE BREAK ---"


def _is_grass_raw_sku(sku: str) -> bool:
    """GRS SKUs are standalone raw ingredient items — excluded from MRP ingredient totals."""
    return "GRS" in sku.upper()


def _is_special_ingredient_sku(sku: str) -> bool:
    """
    NOG / Eggnog / seasonal SKUs have hardcoded ingredient splits in MASTER.
    Do NOT apply Pearson Square to these — use MASTER offsets 21/22/23 instead.
    """
    upper = sku.upper()
    return any(k in upper for k in _SPECIAL_INGREDIENT_KEYWORDS)


# ---------------------------------------------------------------------------
# Pearson Square fallback
# ---------------------------------------------------------------------------

def _pearson_split(
    target_fat: float,
    whole_fat: float,
    skim_fat: float,
    cream_fat: float,
) -> Tuple[float, float, float]:
    """
    Returns (raw_pct, skim_pct, cream_pct).
    Used only as a fallback when Excel's Q/R/S return None (stale formula cache).
    Replicates the Excel Pearson Square formula exactly.

    target_fat <= whole_fat  → standard blend (raw + skim, no cream)
    target_fat >  whole_fat  → cream blend (cream + skim, no raw)
    """
    if target_fat <= whole_fat:
        raw_pct   = (target_fat - skim_fat) / (whole_fat - skim_fat)
        skim_pct  = 1.0 - raw_pct
        cream_pct = 0.0
    else:
        cream_pct = (target_fat - whole_fat) / (cream_fat - whole_fat)
        skim_pct  = 1.0 - cream_pct
        raw_pct   = 0.0
    return raw_pct, skim_pct, cream_pct


# ---------------------------------------------------------------------------
# Validation helpers
# ---------------------------------------------------------------------------

def _validate_row(row: Dict[str, Any], row_idx: int, sheet_name: str) -> Optional[str]:
    """
    Check a parsed schedule row for signs of stale formula cache.
    Returns a warning string if issues are found, None if clean.
    Break rows are skipped — they have no timeline data by design.
    """
    if row.get("isBreakRow"):
        return None

    sku   = row.get("sku", "")
    start = row.get("startHr", 0)
    end   = row.get("endHr", 0)

    if start == 0 and end == 0 and row.get("seq", 1) != 1:
        return (
            f"Sheet '{sheet_name}' Row {row_idx} (SKU: '{sku}'): startHr=0, endHr=0 "
            f"for a non-first row — likely stale formula cache. "
            f"Save the Excel file with full recalculation (Ctrl+Alt+Shift+F9) to fix."
        )

    if row.get("startTime") is None or row.get("endTime") is None:
        return (
            f"Sheet '{sheet_name}' Row {row_idx} (SKU: '{sku}'): startTime or endTime is None — "
            f"formula cell returned no cached value."
        )

    return None


# ---------------------------------------------------------------------------
# Raw material source parser (col Z)
# ---------------------------------------------------------------------------

def _parse_material_sources(z_val) -> Dict[str, Any]:
    """
    Parse col Z string to identify which material code/name corresponds
    to each ingredient type (raw, skim, cream).

    Examples:
      "O.V. RAW (RF000049) & O.V. SKIM (SF000760)"
        → rawCode=RF000049, rawName="O.V. RAW", skimCode=SF000760, ...
      "O.V. Grass raw (SF000881)"
        → rawCode=SF000881, rawName="O.V. Grass raw"
      "O.V. SKIM (SF000760) & O.V. Cream (SF000156)"
        → skimCode=SF000760, creamCode=SF000156
    """
    out: Dict[str, Any] = {
        "rawCode": None,  "rawName": None,
        "skimCode": None, "skimName": None,
        "creamCode": None, "creamName": None,
    }
    if not z_val:
        return out
    for part in str(z_val).split("&"):
        m = re.search(r'(.+?)\s*\(([A-Z0-9]+)\)', part.strip())
        if not m:
            continue
        name = m.group(1).strip()
        code = m.group(2).strip()
        upper = name.upper()
        if "SKIM" in upper:
            out["skimCode"] = code
            out["skimName"] = name
        elif "CREAM" in upper:
            out["creamCode"] = code
            out["creamName"] = name
        elif "RAW" in upper:  # covers both "RAW" and "Grass raw"
            out["rawCode"] = code
            out["rawName"] = name
    return out


# ---------------------------------------------------------------------------
# MASTER — SKU database reader
# ---------------------------------------------------------------------------

def read_sku_database(ws) -> Dict[str, Dict[str, Any]]:
    """
    Read SKU database from MASTER sheet rows 18-58, cols B:Y.
    Returns a dict keyed by UPPERCASE SKU name for case-insensitive lookup.

    Offsets from col B (1-based; openpyxl col B = column index 2):
      Offset  1 = SKU Name         (col B)
      Offset  9 = Fat %            (col J)
      Offset 13 = Run Rate         (col N)
      Offset 14 = Liquifier Req    (col O)  "YES" or blank
      Offset 19 = Gal / Case       (col T)
      Offset 20 = Lbs / Gal        (col U)
      Offset 21 = Raw %            (col V)  Pearson-derived (eggnog = hardcoded)
      Offset 22 = Skim %           (col W)
      Offset 23 = Cream %          (col X)
    """
    sku_db: Dict[str, Dict[str, Any]] = {}
    BASE_COL = 2  # openpyxl col B is index 2

    for row_idx in range(18, 59):
        raw_name = ws.cell(row_idx, BASE_COL).value
        if not raw_name:
            continue
        sku_name = str(raw_name).strip()
        if not sku_name:
            continue

        def _get(offset: int):
            return ws.cell(row_idx, BASE_COL + offset - 1).value

        sources = _parse_material_sources(_get(25))  # col Z
        sku_db[sku_name.upper()] = {
            "name":              sku_name,
            "fatPct":            _get(9),
            "runRate":           _get(13),
            "liquifierRequired": str(_get(14) or "").strip().upper() == "YES",
            "galPerCase":        _get(19),
            "lbsPerGal":         _get(20),
            "rawPct":            _get(21),
            "skimPct":           _get(22),
            "creamPct":          _get(23),
            "rawMaterialSources": str(_get(25) or "").strip(),
            **sources,
        }

    return sku_db


# ---------------------------------------------------------------------------
# MASTER sheet reader
# ---------------------------------------------------------------------------

def read_master_sheet(wb: openpyxl.Workbook) -> Dict[str, Any]:
    """
    Read configuration constants from MASTER sheet.

    V4 layout — 2-column (col B = Line 1, col C = Line 2):
      Row  2: B2 = Line 1 schedule start datetime
              C2 = Line 2 schedule start datetime  ← NOW SEPARATE
      Row  3: B3 = Max run hours
      Row  5: B5 = Run rate (cases/hr)
      Row  6: B6 = CIP duration Line 1 (hrs)
              C6 = CIP duration Line 2 (hrs)
      Row 10: B10 = Whole milk fat %
      Row 11: B11 = Skim fat %
      Row 12: B12 = Cream fat %
      Rows 18–58: SKU database (B:Y)
    """
    ws = wb["MASTER"]

    def _parse_dt(cell_val, label: str) -> datetime:
        if isinstance(cell_val, datetime):
            return cell_val.replace(tzinfo=None)
        logger.warning(
            f"MASTER!{label} is not a datetime — got {type(cell_val).__name__}: {cell_val!r}. "
            f"Falling back to midnight today."
        )
        return datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    l1_start   = _parse_dt(ws["B2"].value, "B2")
    c2_raw     = ws["C2"].value
    l2_start   = _parse_dt(c2_raw, "C2") if c2_raw is not None else l1_start

    sku_db = read_sku_database(ws)

    return {
        # Per-line schedule starts (V4)
        "scheduleStartLine1":  l1_start.isoformat(),
        "scheduleStartLine2":  l2_start.isoformat(),
        # Legacy key — preserved for backward compat with existing consumers
        "scheduleStart":       l1_start.isoformat(),
        # Parameters
        "maxRunHours":         ws["B3"].value or 40,
        "runRate":             ws["B5"].value or 120,
        "cipDurationLine1":    ws["B6"].value or 6,
        "cipDurationLine2":    ws["C6"].value or 6,
        "fatWholeMilkPct":     ws["B10"].value or 4.3,
        "fatSkimPct":          ws["B11"].value or 0.05,
        "fatCreamPct":         ws["B12"].value or 36.0,
        # SKU reference database (used for liquifier detection & ingredient fallback)
        "skuDatabase":         sku_db,
    }


# ---------------------------------------------------------------------------
# Schedule sheet reader
# ---------------------------------------------------------------------------

def read_schedule_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    schedule_start: datetime,
    line_number: int,
    master_params: Optional[Dict[str, Any]] = None,
) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Read schedule rows from a LINE sheet (V4 column layout).

    Column mapping (1-indexed):
      A=1  Seq #           B=2  SKU Name         C=3  Target Fat %
      D=4  Cases           E=5  Run Duration hrs  F=6  Calc Start
      G=7  Calc End        H=8  Override Start    I=9  Override Duration
      J=10 Override End    K=11 Active Start hr   L=12 Active End hr
      M=13 Start Time      N=14 End Time          O=15 Gallons
      P=16 Pounds          Q=17 Raw Req (lbs)     R=18 Skim Req (lbs)
      S=19 Cream Req (lbs) T=20 Recipe Label      U=21 Notes
      V=22 Segment(Line2)  W=23 Segment(Line1)

    Break row timing (Option A):
      On a break row, col H (Override Start) may contain a datetime.
      If so, it is the restart anchor for the next segment — stored as
      segmentStartTime. Excel already recalculated K/L/M/N using this anchor,
      so no timeline math is needed here.

    Segment detection (dual signal):
      1. Segment col W/V = None  → structural indicator of a break/divider row
      2. Col B == "--- SCHEDULE BREAK ---" → explicit isBreakRow flag
      Both signals must agree on break rows; production rows always have segment != None.
    """
    ws = wb[sheet_name]
    rows: List[Dict[str, Any]] = []
    warnings: List[str] = []

    SEGMENT_COL = 23 if line_number == 1 else 22  # W for Line 1, V for Line 2

    # Pearson / ingredient fallback params from master
    whole_fat = (master_params or {}).get("fatWholeMilkPct", 4.3)
    skim_fat  = (master_params or {}).get("fatSkimPct",      0.05)
    cream_fat = (master_params or {}).get("fatCreamPct",     36.0)
    sku_db    = (master_params or {}).get("skuDatabase",     {})

    for row_idx in range(2, ws.max_row + 1):
        sku_raw = ws.cell(row_idx, 2).value
        if not sku_raw:
            continue

        sku = str(sku_raw).strip()

        # --- Row classification ---
        is_break    = _is_break_row(sku)
        is_cip      = not is_break and "CIP" in sku.upper()
        is_flush    = not is_break and ("WATER" in sku.upper() or "FLUSH" in sku.upper())
        is_downtime = not is_break and "DOWNTIME" in sku.upper()
        is_grass    = not is_break and _is_grass_raw_sku(sku)

        # --- Segment (dual signal) ---
        segment_raw = ws.cell(row_idx, SEGMENT_COL).value
        segment: Optional[int] = None
        if not is_break and segment_raw is not None:
            try:
                segment = int(segment_raw)
            except (ValueError, TypeError):
                pass

        # --- Time columns ---
        active_start_hr = ws.cell(row_idx, 11).value  # K
        active_end_hr   = ws.cell(row_idx, 12).value  # L
        start_time_raw  = ws.cell(row_idx, 13).value  # M
        end_time_raw    = ws.cell(row_idx, 14).value  # N

        if not is_break:
            if active_start_hr is None:
                warnings.append(
                    f"Sheet '{sheet_name}' row {row_idx} ('{sku}'): "
                    f"Column K (Active Start Hr) is None — formula not cached."
                )
                active_start_hr = 0.0
            if active_end_hr is None:
                warnings.append(
                    f"Sheet '{sheet_name}' row {row_idx} ('{sku}'): "
                    f"Column L (Active End Hr) is None — formula not cached."
                )
                active_end_hr = 0.0

        def _to_iso(dt_val, fallback_hrs) -> Optional[str]:
            if is_break:
                return None
            if isinstance(dt_val, datetime):
                return dt_val.replace(tzinfo=None).isoformat()
            return (schedule_start + timedelta(hours=float(fallback_hrs or 0))).isoformat()

        start_time_iso = _to_iso(start_time_raw, active_start_hr)
        end_time_iso   = _to_iso(end_time_raw,   active_end_hr)

        # --- Override / segment-restart columns (H, I, J) ---
        h_val             = ws.cell(row_idx, 8).value   # H — Override Start
        override_duration = ws.cell(row_idx, 9).value   # I
        override_end      = ws.cell(row_idx, 10).value  # J

        # Option A: on a break row, a datetime in H is the next segment's restart time.
        # On a production row, H is a regular schedule override (hours or datetime).
        segment_start_time: Optional[str] = None
        has_override = False

        if is_break:
            if isinstance(h_val, datetime):
                segment_start_time = h_val.replace(tzinfo=None).isoformat()
        else:
            has_override = bool(h_val or override_duration or override_end)

        # --- Material columns (Q/R/S) ---
        gallons       = ws.cell(row_idx, 15).value  # O
        pounds        = ws.cell(row_idx, 16).value  # P
        raw_req_lbs   = ws.cell(row_idx, 17).value  # Q  Raw milk requirement (renamed from High Fat)
        skim_req_lbs  = ws.cell(row_idx, 18).value  # R  Skim requirement
        cream_req_lbs = ws.cell(row_idx, 19).value  # S  Cream requirement (V4 new)

        # Pearson fallback: compute ingredient lbs if Excel returned None for Q/R/S.
        # Skipped for: break rows, GRS raw SKUs, utility rows (CIP/flush/downtime).
        if (
            not is_break
            and not is_grass
            and not is_cip
            and not is_flush
            and not is_downtime
            and pounds
            and (raw_req_lbs is None or skim_req_lbs is None or cream_req_lbs is None)
        ):
            target_fat = ws.cell(row_idx, 3).value  # C
            if target_fat is not None:
                # NOG/PUMPKIN/MOCHA/CINN: use MASTER-stored percentages (not Pearson)
                sku_entry = sku_db.get(sku.upper(), {})
                if _is_special_ingredient_sku(sku) and sku_entry:
                    raw_pct   = float(sku_entry.get("rawPct")   or 0.0)
                    skim_pct  = float(sku_entry.get("skimPct")  or 0.0)
                    cream_pct = float(sku_entry.get("creamPct") or 0.0)
                else:
                    raw_pct, skim_pct, cream_pct = _pearson_split(
                        float(target_fat), whole_fat, skim_fat, cream_fat
                    )
                lbs_val = float(pounds)
                if raw_req_lbs is None:
                    raw_req_lbs = lbs_val * raw_pct
                    warnings.append(
                        f"Sheet '{sheet_name}' row {row_idx} ('{sku}'): "
                        f"Q (Raw Req) is None — Pearson fallback applied."
                    )
                if skim_req_lbs is None:
                    skim_req_lbs = lbs_val * skim_pct
                    warnings.append(
                        f"Sheet '{sheet_name}' row {row_idx} ('{sku}'): "
                        f"R (Skim Req) is None — Pearson fallback applied."
                    )
                if cream_req_lbs is None:
                    cream_req_lbs = lbs_val * cream_pct

        row = {
            "seq":              ws.cell(row_idx, 1).value,
            "sku":              sku,
            "targetFatPct":     ws.cell(row_idx, 3).value,
            "cases":            ws.cell(row_idx, 4).value or 0,
            "runDurationHrs":   ws.cell(row_idx, 5).value or 0.0,
            "startHr":          float(active_start_hr or 0),
            "endHr":            float(active_end_hr   or 0),
            "startTime":        start_time_iso,
            "endTime":          end_time_iso,
            "recipeLabel":      ws.cell(row_idx, 20).value,    # T
            "notes":            ws.cell(row_idx, 21).value,    # U
            "segment":          segment,
            # Break row fields
            "isBreakRow":       is_break,
            "segmentStartTime": segment_start_time,            # Option A — datetime in H on break row
            # Row type flags
            "isCIP":            is_cip,
            "isFlush":          is_flush,
            "isDowntime":       is_downtime,
            "isGrassRaw":       is_grass,
            "hasOverride":      has_override,
            # Quantities
            "gallons":          gallons       or 0,
            "pounds":           pounds        or 0,
            "rawReqLbs":        raw_req_lbs   or 0,
            "skimReqLbs":       skim_req_lbs  or 0,
            "creamReqLbs":      cream_req_lbs or 0,
        }

        if not is_break:
            warning = _validate_row(row, row_idx, sheet_name)
            if warning:
                warnings.append(warning)

        rows.append(row)

    return rows, warnings


# ---------------------------------------------------------------------------
# Aggregation helpers
# ---------------------------------------------------------------------------

def _mrp_rows(rows: List[Dict], segment: Optional[int] = None) -> List[Dict]:
    """
    Filter to rows that contribute to ingredient MRP totals.

    Excluded:
      - Schedule break rows (isBreakRow)
      - GRS raw SKUs (standalone raw ingredients, not blend products)
      - CIP / flush / downtime utility rows
    Segment filter applied when segment is not None.
    """
    filtered = [
        r for r in rows
        if not r.get("isBreakRow")
        and not r.get("isGrassRaw")
        and not r.get("isCIP")
        and not r.get("isFlush")
        and not r.get("isDowntime")
    ]
    if segment is not None:
        filtered = [r for r in filtered if r.get("segment") == segment]
    return filtered


# ---------------------------------------------------------------------------
# calculate_materials — three-ingredient MRP, segment-aware
# ---------------------------------------------------------------------------

def calculate_materials(
    line1_rows: List[Dict],
    line2_rows: List[Dict],
    segment: Optional[int] = None,
    sku_db: Optional[Dict] = None,
) -> Dict[str, Any]:
    """
    Aggregate Raw / Skim / Cream ingredient totals across both lines.

    segment=None  → aggregate across all segments ("All" view)
    segment=N     → restrict to a single production run ("Run N" view)
    sku_db        → MASTER SKU database for col Z raw material attribution

    Standard MRP rows use Pearson-derived rawReqLbs/skimReqLbs/creamReqLbs,
    attributed to the specific raw material code from col Z.
    GRS rows bypass Pearson — their pounds go directly to their col Z raw code.

    Returns scalar totals (backward-compatible) plus materialBreakdown:
    a list of {name, code, lbs, skus} sorted by lbs descending.
    """
    def _grs_filtered(rows: List[Dict]) -> List[Dict]:
        result = [r for r in rows if r.get("isGrassRaw") and not r.get("isBreakRow")]
        if segment is not None:
            result = [r for r in result if r.get("segment") == segment]
        return result

    def _build_breakdown(prod_rows: List[Dict], grs_rows: List[Dict]) -> List[Dict]:
        """
        Attribute each row's ingredient lbs to the specific raw material
        identified in col Z, keyed by material NAME (user-friendly).
        prod_rows: Pearson-derived rawReqLbs/skimReqLbs/creamReqLbs
        grs_rows:  direct pound attribution (no Pearson)
        """
        # Key by name (user-friendly), store code as metadata
        acc: Dict[str, Dict] = {}

        def _add(name, code, lbs, sku):
            if not lbs or not name:
                return
            if name not in acc:
                acc[name] = {"name": name, "code": code or "", "lbs": 0.0, "skus": {}}
            acc[name]["lbs"] += float(lbs)
            acc[name]["skus"][sku] = acc[name]["skus"].get(sku, 0.0) + float(lbs)

        for r in prod_rows:
            sku = r["sku"]
            entry = (sku_db or {}).get(sku.upper(), {})
            _add(entry.get("rawName"),   entry.get("rawCode"),   r.get("rawReqLbs",   0), sku)
            _add(entry.get("skimName"),  entry.get("skimCode"),  r.get("skimReqLbs",  0), sku)
            _add(entry.get("creamName"), entry.get("creamCode"), r.get("creamReqLbs", 0), sku)

        # GRS: total pounds → raw material directly (col Z raw code, no Pearson)
        for r in grs_rows:
            sku = r["sku"]
            entry = (sku_db or {}).get(sku.upper(), {})
            lbs = float(r.get("pounds") or r.get("rawReqLbs") or 0)
            _add(entry.get("rawName"), entry.get("rawCode"), lbs, sku)

        result = []
        for name, data in acc.items():
            sku_list = sorted(
                [{"sku": k, "lbs": round(v)} for k, v in data["skus"].items() if v > 0],
                key=lambda x: x["lbs"], reverse=True,
            )
            result.append({
                "name": name,
                "code": data["code"],
                "lbs":  round(data["lbs"]),
                "skus": sku_list,
            })
        return sorted(result, key=lambda x: x["lbs"], reverse=True)

    def _sum(rows: List[Dict]) -> Dict[str, float]:
        prod = _mrp_rows(rows, segment)
        return {
            "gallons":     sum(r["gallons"]     for r in prod),
            "pounds":      sum(r["pounds"]      for r in prod),
            "rawReqLbs":   sum(r["rawReqLbs"]   for r in prod),
            "skimReqLbs":  sum(r["skimReqLbs"]  for r in prod),
            "creamReqLbs": sum(r["creamReqLbs"] for r in prod),
        }

    SCALAR_KEYS = ("gallons", "pounds", "rawReqLbs", "skimReqLbs", "creamReqLbs")
    l1 = _sum(line1_rows)
    l2 = _sum(line2_rows)

    all_prod = _mrp_rows(line1_rows, segment) + _mrp_rows(line2_rows, segment)
    all_grs  = _grs_filtered(line1_rows) + _grs_filtered(line2_rows)

    return {
        "segment":           segment,
        "line1":             l1,
        "line2":             l2,
        "combined":          {k: l1[k] + l2[k] for k in SCALAR_KEYS},
        "materialBreakdown": _build_breakdown(all_prod, all_grs),
    }


# ---------------------------------------------------------------------------
# detect_liquifier_conflicts — cross-line, segment-aware
# ---------------------------------------------------------------------------

def detect_liquifier_conflicts(
    line1_rows: List[Dict],
    line2_rows: List[Dict],
    sku_db: Optional[Dict[str, Any]] = None,
    segment: Optional[int] = None,
) -> Dict[str, Any]:
    """
    Detect scheduling conflicts where both lines concurrently require
    the shared liquifier.

    Detection: for every L1 liquifier row × every L2 liquifier row,
    check interval overlap:  L1.start < L2.end  AND  L2.start < L1.end

    Liquifier requirement is read from MASTER SKU database (offset 14).
    Always re-read from MASTER — the list may change across workbook versions.

    segment=None checks across the full schedule.
    segment=N restricts to a single production run.
    """
    def _requires_liquifier(sku: str) -> bool:
        if not sku_db:
            return False
        entry = sku_db.get(sku.upper())
        return bool(entry and entry.get("liquifierRequired"))

    def _liq_rows(rows: List[Dict]) -> List[Dict]:
        result = [
            r for r in rows
            if not r.get("isBreakRow")
            and _requires_liquifier(r.get("sku", ""))
            and r.get("startHr", 0) != r.get("endHr", 0)
        ]
        if segment is not None:
            result = [r for r in result if r.get("segment") == segment]
        return result

    l1_liq = _liq_rows(line1_rows)
    l2_liq = _liq_rows(line2_rows)

    conflicts = []
    for r1 in l1_liq:
        for r2 in l2_liq:
            if r1["startHr"] < r2["endHr"] and r2["startHr"] < r1["endHr"]:
                conflicts.append({
                    "line1Sku":       r1["sku"],
                    "line1Segment":   r1.get("segment"),
                    "line1StartHr":   r1["startHr"],
                    "line1EndHr":     r1["endHr"],
                    "line2Sku":       r2["sku"],
                    "line2Segment":   r2.get("segment"),
                    "line2StartHr":   r2["startHr"],
                    "line2EndHr":     r2["endHr"],
                    "overlapStartHr": max(r1["startHr"], r2["startHr"]),
                    "overlapEndHr":   min(r1["endHr"],   r2["endHr"]),
                })

    return {
        "conflictCount": len(conflicts),
        "conflicts":     conflicts,
        "segment":       segment,
    }


# ---------------------------------------------------------------------------
# calculate_alerts — per-line, segment-aware
# ---------------------------------------------------------------------------

def calculate_alerts(
    line_rows: List[Dict],
    max_run_hours: float,
    planning_horizon_hours: float = 40.0,
    segment: Optional[int] = None,
) -> Dict[str, Any]:
    """
    Compute alert conditions for a single production line.
    segment=None → full schedule; segment=N → single run scope.
    Break rows are excluded from all alert calculations.
    """
    rows = line_rows
    if segment is not None:
        rows = [r for r in rows if r.get("segment") == segment]

    non_break  = [r for r in rows if not r.get("isBreakRow")]
    prod_rows  = [
        r for r in non_break
        if not r.get("isCIP") and not r.get("isFlush") and not r.get("isDowntime")
    ]

    if not non_break:
        return {
            "overlapDetected":        False,
            "exceedsCIPLimit":        False,
            "cipScheduled":           False,
            "activeOverrides":        0,
            "exceedsPlanningHorizon": False,
            "maxHrs":                 0.0,
        }

    overlap_detected = False
    for i in range(len(prod_rows) - 1):
        r1, r2 = prod_rows[i], prod_rows[i + 1]
        # Only flag overlap within the same segment — different segments reset hour 0
        if r1.get("segment") == r2.get("segment") and r1["endHr"] > r2["startHr"]:
            overlap_detected = True
            break

    cip_scheduled    = any(r.get("isCIP") for r in non_break)
    max_hrs          = max((r["endHr"] for r in non_break), default=0.0)
    active_overrides = sum(1 for r in non_break if r.get("hasOverride"))

    return {
        "overlapDetected":        overlap_detected,
        "exceedsCIPLimit":        max_hrs > max_run_hours,
        "cipScheduled":           cip_scheduled,
        "activeOverrides":        active_overrides,
        "exceedsPlanningHorizon": max_hrs > planning_horizon_hours,
        "maxHrs":                 round(max_hrs, 2),
    }


# ---------------------------------------------------------------------------
# build_gantt_data — segments, per-line schedule starts, break dividers
# ---------------------------------------------------------------------------

def build_gantt_data(
    line1_rows: List[Dict],
    line2_rows: List[Dict],
    schedule_start_line1: str,
    schedule_start_line2: str,
    horizon_hours: int = 72,
) -> Dict[str, Any]:
    """
    Build the Gantt chart data structure for both lines.

    Break rows are included as visual dividers (type="break") — they carry
    no startTime/endTime/bar, but carry segmentStartTime if H was populated.

    Each production task includes a `segment` field for frontend filter support.
    The returned `segments` list is the union of all segment numbers across both lines.
    """

    def _get_segments(rows: List[Dict]) -> List[int]:
        seen: set = set()
        result: List[int] = []
        for r in rows:
            s = r.get("segment")
            if s is not None and s not in seen:
                seen.add(s)
                result.append(s)
        return sorted(result)

    def _build_tasks(rows: List[Dict], line_prefix: str) -> List[Dict]:
        tasks = []
        for r in rows:
            if r.get("isBreakRow"):
                tasks.append({
                    "id":               f"{line_prefix}-break-{r.get('seq', 'x')}",
                    "sku":              r["sku"],
                    "type":             "break",
                    "color":            "#334155",
                    "isBreak":          True,
                    "segment":          None,
                    "segmentStartTime": r.get("segmentStartTime"),
                })
                continue

            tasks.append({
                "id":          f"{line_prefix}-{r['seq']}",
                "sku":         r["sku"],
                "cases":       r["cases"],
                "startTime":   r["startTime"],
                "endTime":     r["endTime"],
                "startHr":     r["startHr"],
                "endHr":       r["endHr"],
                "type":        get_task_type(r["sku"]),
                "color":       get_task_color(r["sku"]),
                "hasOverride": r["hasOverride"],
                "gallons":     r["gallons"],
                "segment":     r.get("segment"),
                "recipeLabel": r.get("recipeLabel"),
                "notes":       r.get("notes"),
                "isGrassRaw":  r.get("isGrassRaw", False),
            })
        return tasks

    seg_l1      = _get_segments(line1_rows)
    seg_l2      = _get_segments(line2_rows)
    all_segments = sorted(set(seg_l1) | set(seg_l2))

    return {
        "scheduleStartLine1": schedule_start_line1,
        "scheduleStartLine2": schedule_start_line2,
        "scheduleStart":      schedule_start_line1,   # legacy key
        "horizonHours":       horizon_hours,
        "segments":           all_segments,
        "lines": [
            {
                "id":            "line1",
                "label":         "LINE 1 (EH)",
                "scheduleStart": schedule_start_line1,
                "tasks":         _build_tasks(line1_rows, "L1"),
            },
            {
                "id":            "line2",
                "label":         "LINE 2 (TR7)",
                "scheduleStart": schedule_start_line2,
                "tasks":         _build_tasks(line2_rows, "L2"),
            },
        ],
    }


# ---------------------------------------------------------------------------
# Workbook loaders
# ---------------------------------------------------------------------------

def load_workbook_safe(file_path: Path) -> openpyxl.Workbook:
    """
    Open workbook with data_only=True.

    CRITICAL: read_only=True is intentionally NOT used here.
    openpyxl's read_only mode is incompatible with data_only — when both
    are set, formula cells return None instead of their cached values.

    Windows sharing-violation fix: copy to a temp file first so we never
    hold an OS-level lock on the original .xlsx while Excel is open.
    """
    import shutil
    import tempfile

    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)
        shutil.copy2(file_path, tmp_path)
    except PermissionError:
        raise PermissionError(
            f"Excel file is locked: {file_path}. "
            f"Close Excel or wait for it to finish saving, then retry."
        )
    except Exception as exc:
        raise RuntimeError(f"Failed to copy workbook at {file_path}: {exc}") from exc

    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        return wb
    except Exception as exc:
        raise RuntimeError(f"Failed to open workbook at {file_path}: {exc}") from exc
    finally:
        try:
            tmp_path.unlink()
        except Exception:
            pass


def load_all_from_bytes(file_bytes: bytes):
    """
    Cloud / upload entry point — accepts raw .xlsx bytes.
    Returns (master, line1_rows, line2_rows, all_warnings).
    """
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    master   = read_master_sheet(wb)
    l1_start = datetime.fromisoformat(master["scheduleStartLine1"])
    l2_start = datetime.fromisoformat(master["scheduleStartLine2"])

    line1_rows, l1_warnings = read_schedule_sheet(
        wb, "SCHEDULE - LINE 1 (EH)", l1_start, 1, master
    )
    line2_rows, l2_warnings = read_schedule_sheet(
        wb, "SCHEDULE - LINE 2 (TR7)", l2_start, 2, master
    )

    all_warnings = l1_warnings + l2_warnings
    for w in all_warnings:
        logger.warning(w)
    if not all_warnings:
        logger.info("Data quality check passed — no formula caching issues detected.")

    wb.close()
    return master, line1_rows, line2_rows, all_warnings


def load_all(file_path: Path):
    """
    Disk entry point. Returns (master, line1_rows, line2_rows, all_warnings).
    Called by api_bridge — never call this directly from routes.
    """
    wb = load_workbook_safe(file_path)

    master   = read_master_sheet(wb)
    l1_start = datetime.fromisoformat(master["scheduleStartLine1"])
    l2_start = datetime.fromisoformat(master["scheduleStartLine2"])

    line1_rows, l1_warnings = read_schedule_sheet(
        wb, "SCHEDULE - LINE 1 (EH)", l1_start, 1, master
    )
    line2_rows, l2_warnings = read_schedule_sheet(
        wb, "SCHEDULE - LINE 2 (TR7)", l2_start, 2, master
    )

    all_warnings = l1_warnings + l2_warnings
    for w in all_warnings:
        logger.warning(w)
    if not all_warnings:
        logger.info("Data quality check passed — no formula caching issues detected.")

    wb.close()
    return master, line1_rows, line2_rows, all_warnings
