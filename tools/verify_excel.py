#!/usr/bin/env python3
"""
Phase 2: Link - Verify Excel file connectivity
Minimal handshake to test openpyxl can read the file
"""

import sys
from pathlib import Path

try:
    import openpyxl
    print("✓ openpyxl imported successfully")
except ImportError:
    print("✗ openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

# File path
file_path = Path(__file__).parent.parent / "St Paul Production Tool V2 Trial.xlsx"

print(f"\nAttempting to read: {file_path}")
print(f"File exists: {file_path.exists()}")

if not file_path.exists():
    print("✗ File not found!")
    sys.exit(1)

try:
    # Open workbook
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print(f"✓ Workbook loaded successfully")

    # List sheets
    print(f"\nSheets found ({len(wb.sheetnames)}):")
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"  {i}. {name}")

    # Test reading MASTER sheet B1 (Schedule Start)
    if 'MASTER' in wb.sheetnames:
        ws = wb['MASTER']
        schedule_start = ws['B1'].value
        print(f"\n✓ MASTER!B1 (Schedule Start): {schedule_start}")
        print(f"  Type: {type(schedule_start)}")

    # Test reading first row from LINE 1
    if 'SCHEDULE - LINE 1 (EH)' in wb.sheetnames:
        ws = wb['SCHEDULE - LINE 1 (EH)']
        # Row 2 is first data row (row 1 is headers)
        sku = ws['B2'].value
        cases = ws['D2'].value
        print(f"\n✓ LINE 1 Row 2:")
        print(f"  SKU: {sku}")
        print(f"  Cases: {cases}")

    print("\n✓ Excel connectivity verified!")

except Exception as e:
    print(f"✗ Error reading Excel: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
